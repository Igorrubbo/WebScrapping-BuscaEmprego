from bs4 import BeautifulSoup
from openpyxl import load_workbook
import requests
from datetime import date


palavras_chave = 'Analista BI Business Intelligence Inteligência Dados Ciência'
cidade = 'Atibaia'

# Carregar planilha no openpyxl para inserir dados e para puxar informação de quais links já foram puxados para a planilha
wb = load_workbook(r'C:\Users\Igor\Desktop\Procurar emprego.xlsx')
ws = wb[wb.sheetnames[0]]
lista_vagas = []
coluna_link = ws['D']
for cell in coluna_link:
    lista_vagas.append(cell.value)
print(lista_vagas)

# Data do dia de hoje em formato dd/mm/yy para saber diferenciar as datas que os links foram puxados
today = date.today()
dia_puxado = today.strftime("%d/%m/%Y")

# O número dentro do range(#) é a quantidade de páginas que serão procuradas. Se quiser olhar mais páginas é só mudar esse valor
lista_palavras_chave = palavras_chave.split()
celula = 0
for i in range(15):
    # URL e busca das informações do site
    url = r'https://www.atibaiasp.com.br/banco-de-vagas/page/' + str(i+1) + "/"
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')
    job_list = soup.find(class_='job-list')
    itens = job_list.find_all(class_='single')

    # Criei algumas condições para garantir que a vaga seja em Atibaia, que as palavras chaves estejam no título da vaga e que o link não tenha aparecido em outro momento
    for item in itens:
        local = item.find(class_='job-location meta-item')
        if cidade.upper() in local.text.strip().upper():
            titulo = item.find('div', attrs={'class': 'title'})
            for palavra in lista_palavras_chave:
                if palavra.upper() in (palavra_titulo.upper() for palavra_titulo in titulo.text.strip("\n").split()):
                    link = titulo.find('a').get('href')
                    if link in lista_vagas:
                        continue
                    # Depois de todas as condições checadas as informações são passadas para a planilha de excel
                    else:
                        celula += 1
                        empresa = item.find(
                            'div', attrs={'class': 'company-name'})
                        #print("Título da vaga: " + titulo.text.strip("\n"))
                        ws['A' + str(celula + 1)] = dia_puxado
                        ws['B' + str(celula + 1)] = titulo.text.strip("\n")
                        #print("Empresa contratante: " + empresa.text)
                        ws['C' + str(celula + 1)] = empresa.text
                        # print(link)
                        ws['D' + str(celula + 1)].hyperlink = link
                        # print('-----------------------------------------------------------------\n')


# Salvar e fechar o arquivo
wb.save(filename=r'C:\Users\Igor\Desktop\Procurar emprego.xlsx')
wb.close()
